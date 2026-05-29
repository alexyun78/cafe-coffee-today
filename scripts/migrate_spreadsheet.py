"""스프레드시트(로스팅기록_2026) → DB 마이그레이션.

Google Drive MCP `read_file_content` 가 내보낸 파일(JSON {fileContent})을 파싱하여
suppliers, green_beans, purchases, roasting_logs 테이블에 적재.

내보낸 내용은 여러 탭이 파이프(|) 구분 마크다운 표로 이어붙여져 있다.
표는 헤더 다음 줄의 `:-:` 구분선으로 식별하고, 컬럼은 헤더 이름으로 매핑한다
(행 위치 하드코딩 X — 시트가 늘어나도 안전).

Usage:
    python scripts/migrate_spreadsheet.py <export_file_path> [--reload]

    --reload : 적재 전에 roasting_logs, purchases 테이블을 비운다 (전체 교체).
               (suppliers/green_beans 는 유지하고 없는 항목만 추가)
"""
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db

SUPPLIER_MAP = {
    "레햄코리아": "레햄",
    "커피리브레": "리브레",
    "소펙스코리아": "소펙",
    "커만사": "커만사",
    "커피플랜트": "커플",
    "알마씨엘로": "알마",
    "더드립": "더드립",
}


def load_content(path: str) -> str:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("fileContent", "")


def parse_cells(line: str) -> list:
    cells = [c.strip() for c in line.split("|")]
    if cells and cells[0] == "":
        cells = cells[1:]
    if cells and cells[-1] == "":
        cells = cells[:-1]
    return cells


def is_separator(line: str) -> bool:
    return bool(re.search(r":-+:", line)) and "|" in line


def find_tables(lines: list) -> list:
    """[(header_cells, [row_cells,...]), ...]. 헤더는 :-: 구분선 바로 위 줄."""
    sep_idx = [i for i in range(1, len(lines)) if is_separator(lines[i])]
    headers = [s - 1 for s in sep_idx]
    tables = []
    for k, h in enumerate(headers):
        start = h + 2
        end = headers[k + 1] if k + 1 < len(headers) else len(lines)
        rows = [parse_cells(lines[x]) for x in range(start, end)]
        tables.append((parse_cells(lines[h]), rows))
    return tables


def _norm_header(s: str) -> str:
    return re.sub(r"&#\d+;", " ", s).replace("\\", "").strip()


def find_col(header: list, *keywords, exclude=()) -> int:
    for idx, name in enumerate(header):
        n = _norm_header(name)
        if all(k in n for k in keywords) and not any(e in n for e in exclude):
            return idx
    return -1


def clean_name(raw: str) -> tuple:
    raw = raw.replace("\\[", "[").replace("\\]", "]").strip()
    m = re.match(r"^\[([^\]]+)\]\s*(.+)$", raw)
    if m:
        prefix, name = m.group(1).strip(), m.group(2).strip()
    else:
        prefix, name = "", raw
    cup_match = re.search(r"\(([^)]+)\)$", name)
    cup_notes = cup_match.group(1).strip() if cup_match else ""
    if cup_match:
        name = name[: cup_match.start()].strip()
    return prefix, name, cup_notes


def parse_date(d: str) -> str:
    d = d.replace("\\.", ".").strip()
    m = re.match(r"(\d{4})\.\s*(\d{1,2})\.\s*(\d{1,2})", d)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return d


def parse_number(s: str) -> float:
    s = (s or "").replace(",", "").replace("\\", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0


def ensure_supplier(name: str) -> int:
    short = SUPPLIER_MAP.get(name, name)
    with db.connect() as conn:
        row = conn.execute("SELECT id FROM suppliers WHERE name=?", (name,)).fetchone()
        if row:
            return row["id"]
    return db.create_supplier({"name": name, "short_name": short})


def ensure_green_bean(name, supplier_id, process, grade="", cup_notes="", is_decaf=0) -> int:
    with db.connect() as conn:
        row = conn.execute(
            "SELECT id FROM green_beans WHERE name=? AND supplier_id=? AND process=?",
            (name, supplier_id, process),
        ).fetchone()
        if row:
            if cup_notes:
                conn.execute(
                    "UPDATE green_beans SET cup_notes=? WHERE id=? AND (cup_notes IS NULL OR cup_notes='')",
                    (cup_notes, row["id"]),
                )
            return row["id"]
    return db.create_green_bean({
        "name": name, "supplier_id": supplier_id, "process": process,
        "grade": grade, "cup_notes": cup_notes, "is_decaf": is_decaf,
    })


def _resolve_bean(raw_name, supplier_name, process, grade):
    prefix, bean_name, cup_notes = clean_name(raw_name)
    is_decaf = 1 if ("디카페인" in process or "디카페인" in bean_name) else 0
    sid = ensure_supplier(supplier_name)
    gbid = ensure_green_bean(bean_name, sid, process, grade, cup_notes, is_decaf)
    return gbid


def migrate_roasting(table) -> int:
    header, rows = table
    c_date = find_col(header, "날짜")
    c_name = find_col(header, "생두")
    c_sup = find_col(header, "구매처")
    c_proc = find_col(header, "가공")
    c_grade = find_col(header, "등급")
    c_in = find_col(header, "투입", exclude=("배출", "수분"))
    c_out = find_col(header, "배출")
    count = 0
    for cells in rows:
        if max(c_date, c_name, c_sup, c_proc, c_in, c_out) >= len(cells):
            continue
        if not re.search(r"\d{4}", cells[c_date]):
            continue
        raw_name = cells[c_name].strip()
        supplier_name = cells[c_sup].strip()
        process = cells[c_proc].strip()
        grade = cells[c_grade].strip() if 0 <= c_grade < len(cells) else ""
        input_g = parse_number(cells[c_in])
        output_g = parse_number(cells[c_out])
        if not supplier_name or not raw_name or input_g == 0:
            continue
        gbid = _resolve_bean(raw_name, supplier_name, process, grade)
        db.create_roasting_log({
            "green_bean_id": gbid,
            "roast_date": parse_date(cells[c_date]),
            "input_weight_g": input_g,
            "output_weight_g": output_g if output_g > 0 else None,
        })
        count += 1
    print(f"  로스팅 기록: {count}건 적재")
    return count


def migrate_purchases(table) -> int:
    header, rows = table
    c_date = find_col(header, "구입일")
    if c_date < 0:
        c_date = find_col(header, "날짜")
    c_name = find_col(header, "원두명")
    c_sup = find_col(header, "구입처")
    c_proc = find_col(header, "가공")
    c_grade = find_col(header, "등급")
    c_price = find_col(header, "단가")
    c_qty = find_col(header, "수량")
    c_total = find_col(header, "구매액")
    count = 0
    for cells in rows:
        if max(c_date, c_name, c_sup, c_price, c_qty, c_total) >= len(cells):
            continue
        if not re.search(r"\d{4}", cells[c_date]):
            continue
        raw_name = cells[c_name].strip()
        supplier_name = cells[c_sup].strip()
        process = cells[c_proc].strip() if 0 <= c_proc < len(cells) else ""
        grade = cells[c_grade].strip() if 0 <= c_grade < len(cells) else ""
        unit_price = int(parse_number(cells[c_price]))
        qty = parse_number(cells[c_qty])
        total_raw = int(parse_number(cells[c_total]))
        if not supplier_name or not raw_name or qty == 0:
            continue
        gbid = _resolve_bean(raw_name, supplier_name, process, grade)
        computed = int(qty * unit_price)
        discount = computed - total_raw if (total_raw and total_raw < computed) else 0
        db.create_purchase({
            "green_bean_id": gbid,
            "purchase_date": parse_date(cells[c_date]),
            "quantity_kg": qty,
            "unit_price": unit_price,
            "discount": discount,
        })
        count += 1
    print(f"  구매 기록: {count}건 적재")
    return count


def _xlsx_date(v) -> str:
    """openpyxl 셀 값(datetime/date/str) → 'YYYY-MM-DD'."""
    import datetime
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return parse_date(str(v)) if v is not None else ""


def _cell(row, idx):
    return row[idx].value if idx < len(row) else None


def migrate_roasting_xlsx(ws) -> int:
    """로스팅 탭: 2줄 헤더(투입/배출 병합), 데이터 3행부터.
    A=날짜 B=로스팅생두 C=구매처 D=가공방식 E=등급 J=투입(g) K=배출(g)."""
    count = 0
    for row in ws.iter_rows(min_row=3):
        date_v = _cell(row, 0)
        if not date_v:
            continue
        raw_name = str(_cell(row, 1) or "").strip()
        supplier_name = str(_cell(row, 2) or "").strip()
        process = str(_cell(row, 3) or "").strip()
        grade = str(_cell(row, 4) or "").strip()
        input_g = parse_number(str(_cell(row, 9) if _cell(row, 9) is not None else ""))
        out_v = _cell(row, 10)
        output_g = parse_number(str(out_v)) if out_v is not None else 0
        if not supplier_name or not raw_name or input_g == 0:
            continue
        gbid = _resolve_bean(raw_name, supplier_name, process, grade)
        db.create_roasting_log({
            "green_bean_id": gbid,
            "roast_date": _xlsx_date(date_v),
            "input_weight_g": input_g,
            "output_weight_g": output_g if output_g > 0 else None,
        })
        count += 1
    print(f"  로스팅 기록: {count}건 적재")
    return count


def migrate_purchases_xlsx(ws) -> int:
    """구입 탭: 1줄 헤더, 데이터 2행부터.
    A=구입일 B=원두명 C=구입처 D=가공방식 E=등급 F=단가 G=수량(Kg) I=구매액."""
    count = 0
    for row in ws.iter_rows(min_row=2):
        date_v = _cell(row, 0)
        if not date_v:
            continue
        raw_name = str(_cell(row, 1) or "").strip()
        supplier_name = str(_cell(row, 2) or "").strip()
        process = str(_cell(row, 3) or "").strip()
        grade = str(_cell(row, 4) or "").strip()
        unit_price = int(parse_number(str(_cell(row, 5) if _cell(row, 5) is not None else "")))
        qty = parse_number(str(_cell(row, 6) if _cell(row, 6) is not None else ""))
        total_raw = int(parse_number(str(_cell(row, 8) if _cell(row, 8) is not None else "")))
        if not supplier_name or not raw_name or qty == 0:
            continue
        gbid = _resolve_bean(raw_name, supplier_name, process, grade)
        computed = int(qty * unit_price)
        discount = computed - total_raw if (total_raw and total_raw < computed) else 0
        db.create_purchase({
            "green_bean_id": gbid,
            "purchase_date": _xlsx_date(date_v),
            "quantity_kg": qty,
            "unit_price": unit_price,
            "discount": discount,
        })
        count += 1
    print(f"  구매 기록: {count}건 적재")
    return count


def run_xlsx(path: str, reload: bool):
    """XLSX(download_file_content export)를 직접 읽어 적재.
    read_file_content 텍스트 export 는 큰 시트를 절삭하므로 XLSX 경로를 권장."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if "로스팅" not in wb.sheetnames or "구입" not in wb.sheetnames:
        print(f"필요한 탭이 없습니다. 발견: {wb.sheetnames}"); sys.exit(1)
    roast_ws, buy_ws = wb["로스팅"], wb["구입"]
    print(f"  로스팅 탭: {roast_ws.max_row}행, 구입 탭: {buy_ws.max_row}행")

    db.init_schema()
    if reload:
        with db.connect() as conn:
            conn.execute("DELETE FROM roasting_logs")
            conn.execute("DELETE FROM purchases")
        print("[--reload] roasting_logs, purchases 비움")
    else:
        with db.connect() as conn:
            existing = conn.execute("SELECT COUNT(*) AS c FROM purchases").fetchone()["c"]
        if existing > 0:
            print(f"이미 {existing}건의 구매 기록이 있습니다. --reload 로 전체 교체하세요.")
            return

    print("\n구매 기록 적재 중...")
    p_count = migrate_purchases_xlsx(buy_ws)
    print("\n로스팅 기록 적재 중...")
    r_count = migrate_roasting_xlsx(roast_ws)
    _report(p_count, r_count)


def _report(p_count, r_count):
    print("\n재고 확인 (상위 일부):")
    for item in db.inventory_list()[:15]:
        remaining = item["remaining_kg"]
        label = (f"[{item.get('supplier_short')}] " if item.get("supplier_short") else "") + item["name"]
        status = "OK" if remaining > 5 else "LOW" if remaining > 0 else "ZERO"
        print(f"  [{status}] {label}: {remaining:.1f}kg (구매 {item['purchased_kg']:.1f} - 사용 {item['used_kg']:.1f})")
    print(f"\n완료: 공급업체 {len(db.list_suppliers())}곳, "
          f"생두 {len(db.list_green_beans(True))}종, "
          f"구매 {p_count}건, 로스팅 {r_count}건")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    reload = "--reload" in sys.argv
    if not args:
        print("Usage: python scripts/migrate_spreadsheet.py <export_file|.xlsx> [--reload]")
        sys.exit(1)
    path = args[0]
    if not os.path.exists(path):
        print(f"파일을 찾을 수 없습니다: {path}")
        sys.exit(1)

    print(f"소스: {path}")
    if path.lower().endswith(".xlsx"):
        run_xlsx(path, reload)
        return

    content = load_content(path)
    lines = content.split("\n")
    tables = find_tables(lines)
    print(f"총 {len(lines)}줄, 표 {len(tables)}개 발견")

    roast_tbl = next((t for t in tables if find_col(t[0], "생두") >= 0
                      and find_col(t[0], "배출") >= 0), None)
    purchase_tbl = next((t for t in tables if find_col(t[0], "구매액") >= 0), None)
    if not roast_tbl:
        print("로스팅 표를 찾지 못했습니다."); sys.exit(1)
    if not purchase_tbl:
        print("구매 표를 찾지 못했습니다."); sys.exit(1)
    print(f"  로스팅 표: 컬럼 {len(roast_tbl[0])}, 행 {len(roast_tbl[1])}")
    print(f"  구매 표: 컬럼 {len(purchase_tbl[0])}, 행 {len(purchase_tbl[1])}")

    db.init_schema()

    if reload:
        with db.connect() as conn:
            conn.execute("DELETE FROM roasting_logs")
            conn.execute("DELETE FROM purchases")
        print("[--reload] roasting_logs, purchases 비움")
    else:
        with db.connect() as conn:
            existing = conn.execute("SELECT COUNT(*) AS c FROM purchases").fetchone()["c"]
        if existing > 0:
            print(f"이미 {existing}건의 구매 기록이 있습니다. --reload 로 전체 교체하세요.")
            return

    print("\n구매 기록 적재 중...")
    p_count = migrate_purchases(purchase_tbl)
    print("\n로스팅 기록 적재 중...")
    r_count = migrate_roasting(roast_tbl)

    print("\n재고 확인 (상위 일부):")
    inv = db.inventory_list()
    for item in inv[:15]:
        remaining = item["remaining_kg"]
        label = (f"[{item.get('supplier_short')}] " if item.get("supplier_short") else "") + item["name"]
        status = "OK" if remaining > 5 else "LOW" if remaining > 0 else "ZERO"
        print(f"  [{status}] {label}: {remaining:.1f}kg (구매 {item['purchased_kg']:.1f} - 사용 {item['used_kg']:.1f})")

    print(f"\n완료: 공급업체 {len(db.list_suppliers())}곳, "
          f"생두 {len(db.list_green_beans(True))}종, "
          f"구매 {p_count}건, 로스팅 {r_count}건")


if __name__ == "__main__":
    main()
